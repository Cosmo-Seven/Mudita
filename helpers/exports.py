from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Q, DateField, DateTimeField
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import os


def filter_queryset_for_export(
    queryset,
    search="",
    date="",
    date_range="",
    date_field="created_at",
    search_fields=None,
):
    # ---------- Search ----------
    if search and search_fields:
        q_obj = Q()
        for field in search_fields:
            q_obj |= Q(**{f"{field}__icontains": search})
        queryset = queryset.filter(q_obj)

    # ---------- Determine field type ----------
    field = queryset.model._meta.get_field(date_field)
    is_datetime = isinstance(field, DateTimeField)
    is_date = isinstance(field, DateField)

    # ---------- Single date ----------
    if date and (is_date or is_datetime):
        try:
            date_obj = datetime.strptime(date, "%Y-%m-%d").date()
            if is_datetime:
                queryset = queryset.filter(**{f"{date_field}__date": date_obj})
            else:
                queryset = queryset.filter(**{f"{date_field}": date_obj})
        except ValueError:
            pass

    # ---------- Date range ----------
    if date_range and "to" in date_range and (is_date or is_datetime):
        try:
            start_date, end_date = [d.strip() for d in date_range.split("to")]
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            if is_datetime:
                queryset = queryset.filter(
                    **{f"{date_field}__date__range": [start_date, end_date]}
                )
            else:
                queryset = queryset.filter(
                    **{f"{date_field}__range": [start_date, end_date]}
                )
        except ValueError:
            pass

    return queryset


def export_to_pdf(
    filename,
    columns,
    rows,
    landscape_mode=False,
    font_name="Helvetica",
    myanmar_font_path=None,
    title=None,
):
    buffer = BytesIO()
    pagesize = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )
    elements = []
    styles = getSampleStyleSheet()

    # ✅ Register Myanmar Font
    if myanmar_font_path and os.path.exists(myanmar_font_path):
        pdfmetrics.registerFont(TTFont("Pyidaungsu", myanmar_font_path))
        font_name = "Pyidaungsu"

    # ✅ Default Paragraph Style
    wrap_style = ParagraphStyle(
        name="WrapStyle",
        fontName=font_name,
        fontSize=9,
        leading=12,
        wordWrap="CJK",
    )

    # ✅ Title
    if title:
        elements.append(
            Paragraph(
                title,
                ParagraphStyle(
                    name="Title",
                    fontName=font_name,
                    fontSize=16,
                    leading=20,
                    alignment=1,  # center
                ),
            )
        )
        elements.append(Spacer(1, 8))

    # ✅ Table Data
    data = [columns]
    for row in rows:
        data.append([Paragraph(str(cell or ""), wrap_style) for cell in row])

    table = Table(data, colWidths=[None] * len(columns))
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F26522")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_to_excel(filename, columns, rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    sheet.append(columns)

    for row in rows:
        sheet.append(row)

    for col in range(1, len(columns) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
